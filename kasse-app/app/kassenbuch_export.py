"""
Kassenbuch-Export: PDF, XLSX, CSV.
Abhängigkeiten: reportlab, openpyxl  (pip install reportlab openpyxl)
"""
from __future__ import annotations
import io
import csv as _csv
from datetime import datetime


def _euro(cent: int) -> str:
    return f"{cent / 100:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")


def _vorzeichen(cent: int) -> str:
    """Formatiert Betrag mit explizitem +/- Vorzeichen."""
    if cent >= 0:
        return f"+{_euro(cent)}"
    return f"−{_euro(-cent)}"


# ── PDF ────────────────────────────────────────────────────────────────────────

def als_pdf(daten: dict, firma_name: str) -> bytes:
    """Erzeugt PDF-Bytes für das Kassenbuch eines Monats."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, PageBreak
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    buf = io.BytesIO()
    PAGE_W, PAGE_H = A4
    MARGIN = 1.5 * cm

    titel   = f"Kassenbuch – {daten['monat_name']} {daten['jahr']}"
    anfang  = daten['anfangssaldo']
    ende    = daten['endsaldo']
    jetzt   = datetime.now().strftime('%d.%m.%Y %H:%M')

    # Styles
    normal  = ParagraphStyle('n',  fontName='Helvetica',      fontSize=9,  leading=11)
    bold    = ParagraphStyle('b',  fontName='Helvetica-Bold',  fontSize=9,  leading=11)
    small   = ParagraphStyle('s',  fontName='Helvetica',      fontSize=7,  leading=9, textColor=colors.grey)
    header_s = ParagraphStyle('h', fontName='Helvetica-Bold',  fontSize=13, leading=16)
    sub_s    = ParagraphStyle('u', fontName='Helvetica',       fontSize=9,  leading=11, textColor=colors.grey)
    right_s  = ParagraphStyle('r', fontName='Helvetica',       fontSize=9,  leading=11, alignment=TA_RIGHT)
    right_bold = ParagraphStyle('rb', fontName='Helvetica-Bold', fontSize=9, leading=11, alignment=TA_RIGHT)

    COL_WIDTHS = [2.4*cm, 2.4*cm, 8.2*cm, 2.4*cm, 2.2*cm]
    HDR_COLOR  = colors.HexColor('#e8f4e8')
    ALT_COLOR  = colors.HexColor('#f9f9f9')
    LINE_COLOR = colors.HexColor('#cccccc')

    # PageNumCanvas: ermöglicht "Seite X von Y"
    from reportlab.pdfgen.canvas import Canvas as _RLCanvas

    class _PageNumCanvas(_RLCanvas):
        """Puffert alle Seiten und schreibt am Ende 'Seite X von Y'."""
        def __init__(self, *args, **kwargs):
            _RLCanvas.__init__(self, *args, **kwargs)
            self._page_states: list = []

        def showPage(self):
            self._page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._page_states)
            for i, state in enumerate(self._page_states, 1):
                self.__dict__.update(state)
                self.setFont('Helvetica', 7)
                self.setFillColor(colors.grey)
                self.drawRightString(
                    PAGE_W - MARGIN, 2.8 * cm - 1.6 * cm,
                    f"Seite {i} von {total}",
                )
                _RLCanvas.showPage(self)
            _RLCanvas.save(self)

    # Seitenaufbau mit Header/Footer
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=3.5*cm, bottomMargin=3.5*cm,
    )

    def header_footer(canvas, doc):
        canvas.saveState()

        # ── Header ──────────────────────────────────────────
        canvas.setFont('Helvetica-Bold', 13)
        canvas.drawString(MARGIN, PAGE_H - 1.5*cm, titel)
        canvas.setFont('Helvetica', 9)
        canvas.setFillColor(colors.grey)
        canvas.drawString(MARGIN, PAGE_H - 2.1*cm, firma_name)
        canvas.setFillColor(colors.black)
        canvas.setStrokeColor(colors.HexColor('#aaaaaa'))
        canvas.setLineWidth(0.5)
        canvas.line(MARGIN, PAGE_H - 2.5*cm, PAGE_W - MARGIN, PAGE_H - 2.5*cm)

        # ── Footer ──────────────────────────────────────────
        y0 = 2.8 * cm
        canvas.setStrokeColor(colors.HexColor('#aaaaaa'))
        canvas.line(MARGIN, y0, PAGE_W - MARGIN, y0)

        canvas.setFont('Helvetica-Bold', 9)
        canvas.drawString(MARGIN, y0 - 0.5*cm, f"Anfangssaldo:  {_euro(anfang)}")
        canvas.drawString(MARGIN, y0 - 1.0*cm, f"Endsaldo:       {_euro(ende)}")

        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(MARGIN, y0 - 1.6*cm, f"Gedruckt: {jetzt}")
        # Seitenzahl wird von _PageNumCanvas in save() geschrieben
        canvas.restoreState()

    # Tabellenkopf
    HEADER_ROW = ['Datum', 'Zu-/Abgang', 'Buchungstext', 'Beleg-Nr.', 'Gegenkonto']

    table_data = [HEADER_ROW]
    row_styles = []

    for i, e in enumerate(daten['eintraege'], start=1):
        dat = e['BUCHUNGSDATUM']
        if hasattr(dat, 'strftime'):
            dat_str = dat.strftime('%d.%m.%Y')
        else:
            dat_str = str(dat)[:10]

        betrag = int(e.get('BETRAG', 0))
        # Align right for amounts
        betrag_p = Paragraph(
            f'<font color="{"green" if betrag >= 0 else "red"}">{_vorzeichen(betrag)}</font>',
            right_s
        )
        text_p = Paragraph(str(e.get('buchungstext_eff', '')), normal)

        table_data.append([
            dat_str,
            betrag_p,
            text_p,
            str(e.get('belegnr_eff', '')),
            str(e.get('gegenkonto_eff', '')),
        ])
        if i % 2 == 0:
            row_styles.append(('BACKGROUND', (0, i), (-1, i), ALT_COLOR))

    # Table
    t = Table(table_data, colWidths=COL_WIDTHS, repeatRows=1)
    base_style = TableStyle([
        # Header
        ('BACKGROUND',    (0, 0), (-1, 0),  HDR_COLOR),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  8),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.HexColor('#336633')),
        ('ROWBACKGROUND', (0, 0), (-1, 0),  HDR_COLOR),
        # Body
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('ALIGN',         (1, 0), (1, -1),  'RIGHT'),
        ('ALIGN',         (3, 0), (4, -1),  'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('GRID',          (0, 0), (-1, -1), 0.3, LINE_COLOR),
        ('LINEBELOW',     (0, 0), (-1, 0),  0.8, colors.HexColor('#aaaaaa')),
    ])
    for s in row_styles:
        base_style.add(*s)
    t.setStyle(base_style)

    story = [
        Spacer(1, 0),
        t,
    ]
    doc.build(story, onFirstPage=header_footer, onLaterPages=header_footer,
              canvasmaker=_PageNumCanvas)
    return buf.getvalue()


# ── XLSX ───────────────────────────────────────────────────────────────────────

def als_xlsx(daten: dict, firma_name: str) -> bytes:
    """Erzeugt XLSX-Bytes für das Kassenbuch eines Monats."""
    from openpyxl import Workbook
    from openpyxl.styles import (Font, Alignment, PatternFill,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    titel = f"Kassenbuch {daten['monat_name']} {daten['jahr']}"
    ws.title = f"{daten['monat_name'][:3]} {daten['jahr']}"

    gruen_fill = PatternFill('solid', fgColor='E8F4E8')
    grau_fill  = PatternFill('solid', fgColor='F5F5F5')
    bold_f  = Font(bold=True)
    head_f  = Font(bold=True, color='336633')
    thin    = Side(border_style='thin', color='CCCCCC')
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    center  = Alignment(horizontal='center', vertical='center')
    right_a = Alignment(horizontal='right',  vertical='center')
    left_a  = Alignment(horizontal='left',   vertical='center')

    # Titel
    ws.merge_cells('A1:E1')
    ws['A1'] = titel
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = firma_name
    ws['A2'].font = Font(italic=True, color='666666', size=9)

    # Kopfzeile
    COLS = ['Datum', 'Zu-/Abgang', 'Buchungstext', 'Beleg-Nr.', 'Gegenkonto']
    for col_i, col_name in enumerate(COLS, start=1):
        c = ws.cell(row=4, column=col_i, value=col_name)
        c.font = head_f
        c.fill = gruen_fill
        c.alignment = center
        c.border = border

    # Daten
    for row_i, e in enumerate(daten['eintraege'], start=5):
        dat = e['BUCHUNGSDATUM']
        if hasattr(dat, 'strftime'):
            dat_val = dat.date() if hasattr(dat, 'date') else dat
        else:
            dat_val = str(dat)[:10]

        betrag_eur = int(e.get('BETRAG', 0)) / 100

        row_data = [
            dat_val,
            betrag_eur,
            str(e.get('buchungstext_eff', '')),
            str(e.get('belegnr_eff', '')),
            str(e.get('gegenkonto_eff', '')),
        ]
        fill = grau_fill if row_i % 2 == 0 else None
        for col_i, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.border = border
            if fill:
                c.fill = fill
            if col_i == 1:  # Datum
                c.number_format = 'DD.MM.YYYY'
                c.alignment = center
            elif col_i == 2:  # Betrag
                c.number_format = '#,##0.00 €'
                c.alignment = right_a
                c.font = Font(color=('006600' if betrag_eur >= 0 else 'CC0000'))
            else:
                c.alignment = left_a

    # Fußzeile
    footer_row = len(daten['eintraege']) + 6
    ws.cell(row=footer_row, column=1, value='Anfangssaldo:').font = bold_f
    c = ws.cell(row=footer_row, column=2, value=daten['anfangssaldo'] / 100)
    c.number_format = '#,##0.00 €'; c.font = bold_f
    ws.cell(row=footer_row+1, column=1, value='Endsaldo:').font = bold_f
    c = ws.cell(row=footer_row+1, column=2, value=daten['endsaldo'] / 100)
    c.number_format = '#,##0.00 €'; c.font = bold_f
    ws.cell(row=footer_row+3, column=1,
            value=f"Gedruckt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ).font = Font(size=8, color='888888')

    # Spaltenbreiten
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 13

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── CSV ────────────────────────────────────────────────────────────────────────

def als_csv(daten: dict) -> str:
    """Erzeugt CSV-String (UTF-8 mit BOM für Excel-Kompatibilität)."""
    out = io.StringIO()
    writer = _csv.writer(out, delimiter=';', quoting=_csv.QUOTE_MINIMAL)

    writer.writerow(['Kassenbuch',
                     f"{daten['monat_name']} {daten['jahr']}"])
    writer.writerow(['Anfangssaldo', _euro(daten['anfangssaldo'])])
    writer.writerow([])
    writer.writerow(['Datum', 'Zu-/Abgang (EUR)', 'Buchungstext', 'Beleg-Nr.', 'Gegenkonto'])

    for e in daten['eintraege']:
        dat = e['BUCHUNGSDATUM']
        dat_str = dat.strftime('%d.%m.%Y') if hasattr(dat, 'strftime') else str(dat)[:10]
        betrag_str = f"{int(e.get('BETRAG', 0)) / 100:.2f}".replace('.', ',')
        writer.writerow([
            dat_str,
            betrag_str,
            str(e.get('buchungstext_eff', '')),
            str(e.get('belegnr_eff', '')),
            str(e.get('gegenkonto_eff', '')),
        ])

    writer.writerow([])
    writer.writerow(['Endsaldo', _euro(daten['endsaldo'])])
    writer.writerow(['Gedruckt', datetime.now().strftime('%d.%m.%Y %H:%M')])

    return '\ufeff' + out.getvalue()  # UTF-8 BOM for Excel
