"""
Fester Parser für das „Kramer"-Lieferantenkatalog-Excel.

Layout (live-Beispiel „Aufstellung Kramer Gesamt …xlsx", 2026-05-17):
* 1 Blatt je Marke (Sheet-Name = Marke, z. B. „Wurm").
* Zeile 1 (0-idx 0): Vor-Header (ignoriert).
* Zeile 2 (0-idx 1): Spalten-Header — Texte mit Zahlen-Präfixen wie
  ``5- Kategorie``, ``2 Name``, ``EK NETTO``. Wir mappen über den
  normalisierten Header-NAMEN (Präfix/Whitespace entfernt), nicht
  über feste Spaltenindizes → tolerant ggü. Spalten-Umordnung
  innerhalb dieses Kramer-Formats.
* Zeile 3+ : Artikelzeilen. Ohne Artikel-Nr → übersprungen.

Bewusst nur dieses eine Format (User-Entscheidung „fester Kramer-
Parser"); weitere Lieferanten-Formate kommen als eigene Parser.
"""
from __future__ import annotations

import hashlib
import re
from typing import Any


def _norm_header(s: Any) -> str:
    """Header-Text normalisieren: führende Zahlen/Sonderzeichen +
    Whitespace weg, klein. ``'5- Kategorie'`` → ``'kategorie'``."""
    t = str(s or '').strip().lower()
    t = re.sub(r'^[\s\d\-.)]+', '', t)        # führende „5- " / „2 "
    t = re.sub(r'[\s:]+$', '', t)
    return re.sub(r'\s+', ' ', t).strip()


# Normalisierter Header → Zielfeld. NUR verlässliche Felder: Kramer
# zerlegt den Freitext-Block „Beschreibung" zusätzlich in viele
# Label/Wert-Spalten (Steuer:/Inhalt:/Pfand: …), wodurch Header wie
# „Kartoneinheit"/„Einheit"/„Handelsklasse"/„Bemerkung" Label statt
# Wert enthalten. Diese werden NICHT als Strukturfeld gemappt — der
# vollständige Original-Block bleibt verlustfrei in ``beschreibung``.
# Bei doppeltem Header (Kramer: 'EAN' Label- + Wert-Spalte) gewinnt
# das LETZTE Vorkommen (= Wertspalte).
_FELD_MAP = {
    'artikel-nr':            'lief_art_nr',
    'artikelnr':             'lief_art_nr',
    'kategorie':             'kategorie',
    'kurzbeschreibung':      'kurzbeschreibung',
    # Spalte „Artikelname" wird IGNORIERT (bei Kramer-Varianten nur
    # ein Fragment, z. B. „- Apfel-Zimt"). Primärer Artikelname =
    # Spalte „2 Name" (User-Entscheidung 2026-05-17).
    'name':                  'artikelname',
    'gebinde':               'gebinde',
    'ek netto':              'ek_netto',
    'beschreibung':          'beschreibung',
    'ust-satz':              'ust_satz',
    'empfohlener preis (lvp)': 'vk_empf',
    'ean':                   'ean',
    'liefermenge min.':      'menge_min',
    'produktbild link':      'bild_url',
}

_DEZIMAL = ('ek_netto', 'ust_satz', 'vk_empf', 'menge_min')


def _zahl(v: Any) -> float | None:
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(' ', ' ')
    m = re.search(r'-?\d{1,3}(?:[.\s]\d{3})*(?:[.,]\d+)?|-?\d+(?:[.,]\d+)?',
                  s)
    if not m:
        return None
    z = m.group(0).replace(' ', '').replace('.', '') \
        if (',' in m.group(0)) else m.group(0).replace(' ', '')
    z = z.replace(',', '.')
    try:
        return float(z)
    except ValueError:
        return None


def _str(v: Any, limit: int) -> str:
    return ('' if v is None else str(v)).strip()[:limit]


def parse_kramer_xlsx(path: str) -> list[dict[str, Any]]:
    """Liest ALLE Blätter (= Marken). Returns Liste von
    ``{'marke': str, 'positionen': [ {feld: wert, ...}, ... ]}``.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    blaetter: list[dict[str, Any]] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        header = rows[1]
        # normalisierter Header → Spaltenindex (letztes Vorkommen
        # gewinnt: Kramers Wert-Spalte schlägt die Label-Spalte).
        spalte: dict[str, int] = {}
        for idx, h in enumerate(header):
            feld = _FELD_MAP.get(_norm_header(h))
            if feld:
                spalte[feld] = idx
        if 'lief_art_nr' not in spalte:
            continue  # ohne Artikel-Nr-Spalte kein verwertbares Blatt

        positionen: list[dict[str, Any]] = []
        for r in rows[2:]:
            if not r:
                continue
            pos: dict[str, Any] = {}
            for feld, idx in spalte.items():
                val = r[idx] if idx < len(r) else None
                if feld in _DEZIMAL:
                    pos[feld] = _zahl(val)
                elif feld == 'lief_art_nr':
                    pos[feld] = _str(val, 60)
                elif feld in ('beschreibung', 'bemerkung'):
                    pos[feld] = ('' if val is None else str(val)).strip()
                else:
                    pos[feld] = _str(val, 255)
            pos['marke'] = sn
            name = (pos.get('name_lang') or pos.get('artikelname')
                    or '').strip()
            # Zeilen OHNE Artikel-Nr trotzdem aufnehmen (User-Wunsch) —
            # aber echte Leer-/Abschnittszeilen (weder Nr noch Name)
            # überspringen.
            if not pos.get('lief_art_nr'):
                if not name:
                    continue
                # Stabiler Ersatz-Schlüssel für UNIQUE (Marke+Name+
                # Gebinde) — deterministisch, damit Re-Import upsertet
                # statt zu duplizieren. Präfix '~' = „ohne Lief-Nr".
                roh = f"{sn}|{name}|{pos.get('gebinde') or ''}"
                pos['lief_art_nr'] = '~' + hashlib.sha1(
                    roh.encode('utf-8')).hexdigest()[:12]
                pos['ohne_liefnr'] = True
            positionen.append(pos)
        if positionen:
            blaetter.append({'marke': sn, 'positionen': positionen})
    return blaetter
